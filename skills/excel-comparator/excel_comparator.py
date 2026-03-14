#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Comparator - 批次比對 Excel 檔案差異
支援內容、格式、樣式比對，生成詳細報告
"""

import sys
import os
from datetime import datetime, date, time
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional
from dataclasses import dataclass, field
import json
import argparse

try:
    from openpyxl import load_workbook
    from openpyxl.cell import Cell
    from openpyxl.styles import Font, Fill, Alignment, Border
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


@dataclass
class CellDifference:
    """儲存格差異資訊"""
    sheet: str
    cell: str
    row: int
    col: int
    diff_type: str  # value, format, style
    field: str  # 具體欄位
    value1: Any
    value2: Any
    description: str = ""


@dataclass
class ComparisonResult:
    """比對結果"""
    file1: str
    file2: str
    timestamp: str
    total_cells_compared: int = 0
    differences: List[CellDifference] = field(default_factory=list)
    sheets_only_in_file1: List[str] = field(default_factory=list)
    sheets_only_in_file2: List[str] = field(default_factory=list)
    summary: Dict[str, int] = field(default_factory=dict)

    def add_difference(self, diff: CellDifference):
        self.differences.append(diff)
        key = f"{diff.diff_type}_{diff.field}"
        self.summary[key] = self.summary.get(key, 0) + 1


class ExcelComparator:
    """Excel 檔案比對器"""

    def __init__(
        self,
        compare_values: bool = True,
        compare_formats: bool = True,
        compare_styles: bool = False,
        tolerance: float = 1e-10,
        ignore_whitespace: bool = True,
        compare_formulas: bool = False
    ):
        self.compare_values = compare_values
        self.compare_formats = compare_formats
        self.compare_styles = compare_styles
        self.tolerance = tolerance
        self.ignore_whitespace = ignore_whitespace
        self.compare_formulas = compare_formulas

    def compare(self, file1: str, file2: str, sheets: Optional[List[str]] = None) -> ComparisonResult:
        """比對兩個 Excel 檔案"""
        result = ComparisonResult(
            file1=file1,
            file2=file2,
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )

        # 載入工作簿
        wb1 = load_workbook(file1, data_only=not self.compare_formulas)
        wb2 = load_workbook(file2, data_only=not self.compare_formulas)

        # 取得工作表清單
        sheets1 = set(wb1.sheetnames)
        sheets2 = set(wb2.sheetnames)

        result.sheets_only_in_file1 = list(sheets1 - sheets2)
        result.sheets_only_in_file2 = list(sheets2 - sheets1)

        # 決定要比對的工作表
        common_sheets = sheets1 & sheets2
        if sheets:
            common_sheets = common_sheets & set(sheets)

        # 比對每個工作表
        for sheet_name in sorted(common_sheets):
            ws1 = wb1[sheet_name]
            ws2 = wb2[sheet_name]
            self._compare_sheets(ws1, ws2, sheet_name, result)

        wb1.close()
        wb2.close()

        return result

    def _compare_sheets(self, ws1, ws2, sheet_name: str, result: ComparisonResult):
        """比對兩個工作表"""
        # 取得最大範圍
        max_row = max(ws1.max_row or 1, ws2.max_row or 1)
        max_col = max(ws1.max_column or 1, ws2.max_column or 1)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell1 = ws1.cell(row=row, column=col)
                cell2 = ws2.cell(row=row, column=col)
                cell_ref = f"{get_column_letter(col)}{row}"

                result.total_cells_compared += 1
                self._compare_cells(cell1, cell2, sheet_name, cell_ref, row, col, result)

    def _compare_cells(
        self,
        cell1: Cell,
        cell2: Cell,
        sheet: str,
        cell_ref: str,
        row: int,
        col: int,
        result: ComparisonResult
    ):
        """比對兩個儲存格"""
        # 比對值
        if self.compare_values:
            self._compare_values(cell1, cell2, sheet, cell_ref, row, col, result)

        # 比對格式
        if self.compare_formats:
            self._compare_formats(cell1, cell2, sheet, cell_ref, row, col, result)

        # 比對樣式
        if self.compare_styles:
            self._compare_styles(cell1, cell2, sheet, cell_ref, row, col, result)

    def _compare_values(
        self,
        cell1: Cell,
        cell2: Cell,
        sheet: str,
        cell_ref: str,
        row: int,
        col: int,
        result: ComparisonResult
    ):
        """比對儲存格值"""
        val1 = cell1.value
        val2 = cell2.value

        # 處理空值
        if val1 is None and val2 is None:
            return
        if val1 is None or val2 is None:
            result.add_difference(CellDifference(
                sheet=sheet, cell=cell_ref, row=row, col=col,
                diff_type="value", field="value",
                value1=self._format_value(val1),
                value2=self._format_value(val2),
                description="One cell is empty"
            ))
            return

        # 處理字串
        if isinstance(val1, str) and isinstance(val2, str):
            str1 = val1.strip() if self.ignore_whitespace else val1
            str2 = val2.strip() if self.ignore_whitespace else val2
            if str1 != str2:
                result.add_difference(CellDifference(
                    sheet=sheet, cell=cell_ref, row=row, col=col,
                    diff_type="value", field="text",
                    value1=val1, value2=val2,
                    description="Text values differ"
                ))
            return

        # 處理數字（包含容差比對）
        if isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
            if abs(float(val1) - float(val2)) > self.tolerance:
                result.add_difference(CellDifference(
                    sheet=sheet, cell=cell_ref, row=row, col=col,
                    diff_type="value", field="number",
                    value1=val1, value2=val2,
                    description=f"Numeric values differ (tolerance: {self.tolerance})"
                ))
            return

        # 處理日期時間
        if isinstance(val1, (datetime, date, time)) or isinstance(val2, (datetime, date, time)):
            if val1 != val2:
                result.add_difference(CellDifference(
                    sheet=sheet, cell=cell_ref, row=row, col=col,
                    diff_type="value", field="datetime",
                    value1=str(val1), value2=str(val2),
                    description="Datetime values differ"
                ))
            return

        # 其他類型直接比對
        if val1 != val2:
            result.add_difference(CellDifference(
                sheet=sheet, cell=cell_ref, row=row, col=col,
                diff_type="value", field="other",
                value1=self._format_value(val1),
                value2=self._format_value(val2),
                description=f"Values differ (types: {type(val1).__name__} vs {type(val2).__name__})"
            ))

    def _compare_formats(
        self,
        cell1: Cell,
        cell2: Cell,
        sheet: str,
        cell_ref: str,
        row: int,
        col: int,
        result: ComparisonResult
    ):
        """比對儲存格格式"""
        fmt1 = cell1.number_format
        fmt2 = cell2.number_format

        if fmt1 != fmt2:
            # 分析格式差異類型
            format_type = self._analyze_format_difference(fmt1, fmt2)
            result.add_difference(CellDifference(
                sheet=sheet, cell=cell_ref, row=row, col=col,
                diff_type="format", field=format_type,
                value1=fmt1, value2=fmt2,
                description=f"Number format differs: '{fmt1}' vs '{fmt2}'"
            ))

    def _analyze_format_difference(self, fmt1: str, fmt2: str) -> str:
        """分析格式差異類型"""
        # 檢查是否為日期格式
        date_indicators = ['y', 'm', 'd', 'h', 's', 'AM', 'PM']
        is_date1 = any(ind in fmt1 for ind in date_indicators)
        is_date2 = any(ind in fmt2 for ind in date_indicators)
        if is_date1 or is_date2:
            return "datetime_format"

        # 檢查是否為百分比格式
        if '%' in fmt1 or '%' in fmt2:
            return "percentage_format"

        # 檢查是否為貨幣格式
        currency_indicators = ['$', '¥', '€', '£', 'NT$']
        if any(c in fmt1 or c in fmt2 for c in currency_indicators):
            return "currency_format"

        # 檢查小數位數差異
        decimal1 = fmt1.count('0') if '.' in fmt1 else 0
        decimal2 = fmt2.count('0') if '.' in fmt2 else 0
        if decimal1 != decimal2:
            return "decimal_places"

        return "number_format"

    def _compare_styles(
        self,
        cell1: Cell,
        cell2: Cell,
        sheet: str,
        cell_ref: str,
        row: int,
        col: int,
        result: ComparisonResult
    ):
        """比對儲存格樣式"""
        # 字型比對
        if cell1.font != cell2.font:
            result.add_difference(CellDifference(
                sheet=sheet, cell=cell_ref, row=row, col=col,
                diff_type="style", field="font",
                value1=self._font_to_str(cell1.font),
                value2=self._font_to_str(cell2.font),
                description="Font style differs"
            ))

        # 對齊比對
        if cell1.alignment != cell2.alignment:
            result.add_difference(CellDifference(
                sheet=sheet, cell=cell_ref, row=row, col=col,
                diff_type="style", field="alignment",
                value1=self._alignment_to_str(cell1.alignment),
                value2=self._alignment_to_str(cell2.alignment),
                description="Alignment differs"
            ))

        # 填充/背景顏色比對
        if cell1.fill != cell2.fill:
            result.add_difference(CellDifference(
                sheet=sheet, cell=cell_ref, row=row, col=col,
                diff_type="style", field="fill",
                value1=str(cell1.fill.fgColor.rgb if cell1.fill.fgColor else "None"),
                value2=str(cell2.fill.fgColor.rgb if cell2.fill.fgColor else "None"),
                description="Fill color differs"
            ))

    def _format_value(self, value: Any) -> str:
        """格式化值為字串"""
        if value is None:
            return "<empty>"
        if isinstance(value, (datetime, date, time)):
            return str(value)
        return repr(value)

    def _font_to_str(self, font: Font) -> str:
        """字型轉字串"""
        return f"{font.name}, {font.size}pt, bold={font.bold}, italic={font.italic}"

    def _alignment_to_str(self, alignment: Alignment) -> str:
        """對齊轉字串"""
        return f"h={alignment.horizontal}, v={alignment.vertical}"


class ReportGenerator:
    """報告生成器"""

    @staticmethod
    def generate_markdown(result: ComparisonResult, output_path: str):
        """生成 Markdown 報告"""
        lines = []
        lines.append("# Excel 比對報告")
        lines.append("")
        lines.append("## 基本資訊")
        lines.append("")
        lines.append(f"- **比對時間**: {result.timestamp}")
        lines.append(f"- **檔案 1**: `{result.file1}`")
        lines.append(f"- **檔案 2**: `{result.file2}`")
        lines.append(f"- **比對儲存格數**: {result.total_cells_compared:,}")
        lines.append(f"- **差異數量**: {len(result.differences):,}")
        lines.append("")

        # 工作表差異
        if result.sheets_only_in_file1 or result.sheets_only_in_file2:
            lines.append("## 工作表差異")
            lines.append("")
            if result.sheets_only_in_file1:
                lines.append(f"- 僅存在於檔案 1: {', '.join(result.sheets_only_in_file1)}")
            if result.sheets_only_in_file2:
                lines.append(f"- 僅存在於檔案 2: {', '.join(result.sheets_only_in_file2)}")
            lines.append("")

        # 摘要統計
        if result.summary:
            lines.append("## 差異摘要")
            lines.append("")
            lines.append("| 類型 | 數量 |")
            lines.append("|------|------|")
            for key, count in sorted(result.summary.items()):
                diff_type, field = key.split("_", 1)
                lines.append(f"| {diff_type} - {field} | {count:,} |")
            lines.append("")

        # 詳細差異（限制數量）
        if result.differences:
            lines.append("## 詳細差異")
            lines.append("")

            # 按工作表分組
            by_sheet: Dict[str, List[CellDifference]] = {}
            for diff in result.differences:
                if diff.sheet not in by_sheet:
                    by_sheet[diff.sheet] = []
                by_sheet[diff.sheet].append(diff)

            for sheet_name, diffs in sorted(by_sheet.items()):
                lines.append(f"### 工作表: {sheet_name}")
                lines.append("")
                lines.append("| 儲存格 | 類型 | 欄位 | 檔案 1 | 檔案 2 | 說明 |")
                lines.append("|--------|------|------|--------|--------|------|")

                # 限制每個工作表顯示 100 筆
                for diff in diffs[:100]:
                    val1 = str(diff.value1)[:30].replace("|", "\\|")
                    val2 = str(diff.value2)[:30].replace("|", "\\|")
                    lines.append(f"| {diff.cell} | {diff.diff_type} | {diff.field} | {val1} | {val2} | {diff.description} |")

                if len(diffs) > 100:
                    lines.append(f"| ... | ... | ... | ... | ... | 還有 {len(diffs) - 100} 筆差異 |")
                lines.append("")

        # 結論
        lines.append("## 結論")
        lines.append("")
        if len(result.differences) == 0:
            lines.append("**兩個檔案完全一致**")
        else:
            lines.append(f"**發現 {len(result.differences):,} 處差異**")
            lines.append("")
            lines.append("請檢視上述差異詳情。")

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        return output_path

    @staticmethod
    def generate_json(result: ComparisonResult, output_path: str):
        """生成 JSON 報告"""
        data = {
            "file1": result.file1,
            "file2": result.file2,
            "timestamp": result.timestamp,
            "total_cells_compared": result.total_cells_compared,
            "total_differences": len(result.differences),
            "sheets_only_in_file1": result.sheets_only_in_file1,
            "sheets_only_in_file2": result.sheets_only_in_file2,
            "summary": result.summary,
            "differences": [
                {
                    "sheet": d.sheet,
                    "cell": d.cell,
                    "row": d.row,
                    "col": d.col,
                    "diff_type": d.diff_type,
                    "field": d.field,
                    "value1": str(d.value1),
                    "value2": str(d.value2),
                    "description": d.description
                }
                for d in result.differences
            ]
        }

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Excel Comparator - 比對兩個 Excel 檔案的差異"
    )
    parser.add_argument("file1", help="第一個 Excel 檔案路徑")
    parser.add_argument("file2", help="第二個 Excel 檔案路徑")
    parser.add_argument("-o", "--output", help="輸出報告路徑", default="comparison_report.md")
    parser.add_argument("--format", choices=["md", "json"], default="md", help="報告格式")
    parser.add_argument("--no-values", action="store_true", help="不比對值")
    parser.add_argument("--no-formats", action="store_true", help="不比對格式")
    parser.add_argument("--styles", action="store_true", help="比對樣式")
    parser.add_argument("--formulas", action="store_true", help="比對公式而非計算值")
    parser.add_argument("--tolerance", type=float, default=1e-10, help="數值比對容差")
    parser.add_argument("--sheets", nargs="+", help="指定要比對的工作表名稱")

    args = parser.parse_args()

    # 檢查檔案存在
    if not os.path.exists(args.file1):
        print(f"Error: File not found: {args.file1}")
        sys.exit(1)
    if not os.path.exists(args.file2):
        print(f"Error: File not found: {args.file2}")
        sys.exit(1)

    # 建立比對器
    comparator = ExcelComparator(
        compare_values=not args.no_values,
        compare_formats=not args.no_formats,
        compare_styles=args.styles,
        tolerance=args.tolerance,
        compare_formulas=args.formulas
    )

    print(f"Comparing Excel files...")
    print(f"  File 1: {args.file1}")
    print(f"  File 2: {args.file2}")
    print()

    # 執行比對
    result = comparator.compare(args.file1, args.file2, args.sheets)

    # 生成報告
    if args.format == "md":
        output_path = ReportGenerator.generate_markdown(result, args.output)
    else:
        output_path = ReportGenerator.generate_json(result, args.output)

    print(f"Comparison complete!")
    print(f"  Total cells compared: {result.total_cells_compared:,}")
    print(f"  Differences found: {len(result.differences):,}")
    print(f"  Report saved to: {output_path}")

    if result.summary:
        print()
        print("Summary by type:")
        for key, count in sorted(result.summary.items()):
            print(f"  - {key}: {count}")

    # 回傳差異數量作為 exit code（0 表示相同）
    sys.exit(0 if len(result.differences) == 0 else 1)


if __name__ == "__main__":
    main()
